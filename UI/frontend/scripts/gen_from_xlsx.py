"""엑셀 원본 → TypeScript 데이터 모듈 생성기.

  입력 — data/ 아래, 팀원이 준 원본이라 읽기만 하고 절대 수정하지 않는다:
    data/반도체 불량 분석 개선안.xlsx   공정 × 결함유형 × 유형①②  →  원인 / 해결 / 개선
    data/불량 대응 log.xlsx             315건 대응 이력 (방향성·계측·재발 횟수·조치)

  출력 (생성물 — 직접 손대지 말고 이 스크립트를 다시 돌릴 것):
    src/domain/causeMatrix.generated.ts
    src/services/historySeed.generated.ts

  사용:  UI/frontend 에서  python scripts/gen_from_xlsx.py
"""
import openpyxl, re, json, sys, io, os, collections, datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
# ROOT = UI/frontend (생성물이 들어갈 곳), REPO = 저장소 루트 (data/가 있는 곳)
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPO = os.path.dirname(os.path.dirname(ROOT))
DATA = os.path.join(REPO, 'data')
IMP = os.path.join(DATA, '반도체 불량 분석 개선안.xlsx')
LOG = os.path.join(DATA, '불량 대응 log.xlsx')

BINS = ['Center', 'Donut', 'Edge-Ring', 'Edge-Loc', 'Loc', 'Random', 'Scratch', 'Near-full']
PROCS = ['Diffusion', 'Deposition', 'Photo', 'Etch', 'Cleaning', 'CMP']
PROC_ID = {p: p.upper() for p in PROCS}


def norm(s):
    return re.sub(r'[ \t]+', ' ', str(s or '')).strip()


def lines(s):
    return [norm(x) for x in str(s or '').split('\n') if norm(x)]


# ─────────────────────────────────────────────── 개선안 파싱
def parse_cause(raw):
    """'명칭: X / 상세 설명 / -불릿… / 웨이퍼맵 형태: Y' → dict"""
    ls = lines(raw)
    name, mech, wafer = '', [], ''
    section = None
    for ln in ls:
        m = re.match(r'^명칭\s*[:：]?\s*(.*)$', ln)
        if m:
            name = m.group(1).strip()
            section = 'name'
            continue
        if re.match(r'^상세\s*설명\s*[:：]?\s*$', ln):
            section = 'mech'
            continue
        m = re.match(r'^웨이퍼맵\s*형태\s*[:：]?\s*(.*)$', ln)
        if m:
            wafer = m.group(1).strip()
            section = 'wafer'
            continue
        body = re.sub(r'^[-–·•]\s*', '', ln).strip()
        if not body:
            continue
        if section == 'wafer':
            wafer = (wafer + ' ' + body).strip() if wafer else body
        elif section in ('mech', 'name'):
            mech.append(body)
    return {'name': name, 'mechanism': mech, 'waferMap': wafer}


def split_actions(raw):
    """'A, B, C' → [A, B, C]. 괄호 안 쉼표는 보호한다."""
    s = norm(str(raw or '').replace('\n', ' '))
    if s in ('', 'X'):
        return []
    out, buf, depth = [], '', 0
    for ch in s:
        if ch in '([':
            depth += 1
        elif ch in ')]':
            depth = max(0, depth - 1)
        if ch == ',' and depth == 0:
            if buf.strip():
                out.append(buf.strip())
            buf = ''
        else:
            buf += ch
    if buf.strip():
        out.append(buf.strip())
    return out


wsI = openpyxl.load_workbook(IMP, data_only=True).worksheets[0]
rowsI = list(wsI.iter_rows(values_only=True))
hdrI = [norm(c) for c in rowsI[0]]
BINCOL = {b: hdrI.index(b) for b in BINS}

cells = {}
proc = None
for r in rowsI[1:]:
    if r[0]:
        proc = norm(r[0])
    label = norm(r[1])
    if not label:
        continue
    m = re.match(r'(.+?)\s*불량 유형\s*([①②])\s*(원인|해결|개선)', label)
    if not m:
        continue
    p, var, kind = m.group(1).strip(), (1 if m.group(2) == '①' else 2), m.group(3)
    for b in BINS:
        cells[(p, var, kind, b)] = r[BINCOL[b]]

entries, exclusions = [], []
for p in PROCS:
    for var in (1, 2):
        for b in BINS:
            raw = cells.get((p, var, '원인', b))
            txt = norm(str(raw or '').replace('\n', ' '))
            if txt in ('', 'X'):
                continue
            if not txt.startswith('명칭'):
                # "이 공정은 이 유형을 유발하지 않는다"는 근거 서술
                exclusions.append({'process': PROC_ID[p], 'pattern': b, 'reason': norm(str(raw))})
                continue
            c = parse_cause(raw)
            entries.append({
                'id': f"{p.lower()}-{b.lower().replace('-', '')}-{var}",
                'process': PROC_ID[p], 'pattern': b, 'variant': var,
                'name': c['name'], 'mechanism': c['mechanism'], 'waferMap': c['waferMap'],
                'resolution': split_actions(cells.get((p, var, '해결', b))),
                'improvement': split_actions(cells.get((p, var, '개선', b))),
            })

print(f"개선안: 원인 {len(entries)}건, 유발근거 서술 {len(exclusions)}건")

# ─────────────────────────────────────────────── 로그 파싱
wsL = openpyxl.load_workbook(LOG, data_only=True).worksheets[0]
rowsL = list(wsL.iter_rows(values_only=True))
hi = next(i for i, r in enumerate(rowsL) if norm(r[0]) == '관리번호')
HL = [norm(c) for c in rowsL[hi]]
recs = [{HL[i]: r[i] for i in range(len(HL))} for r in rowsL[hi + 1:] if r[0]]
print(f"로그: {len(recs)}건 / 관리번호 {len({norm(r['관리번호']) for r in recs})}종")

by_id = collections.OrderedDict()
for r in recs:
    by_id.setdefault(norm(r['관리번호']), []).append(r)


def clock_of(s):
    s = norm(s)
    if s == '해당 없음':
        return None
    if s == '장비 구조 의존':
        return {'kind': 'layout', 'label': s}
    m = re.match(r'^(\d{1,2})시→(\d{1,2})시 방향$', s)
    if m:
        return {'kind': 'vector', 'from': int(m.group(1)), 'to': int(m.group(2)), 'label': s}
    m = re.match(r'^(\d{1,2})시 방향$', s)
    if m:
        return {'kind': 'fixed', 'clock': int(m.group(1)), 'label': s}
    return {'kind': 'layout', 'label': s}


def tokens(s):
    return set(re.findall(r'[A-Za-z가-힣]{2,}', norm(s)))


# 로그 관리번호 → 개선안 원인 매칭 (명칭 포함 또는 토큰 최대 겹침)
by_cell = collections.defaultdict(list)
for e in entries:
    by_cell[(e['process'], e['pattern'])].append(e)

link = {}
for mid, rs in by_id.items():
    r0 = rs[0]
    key = (PROC_ID[norm(r0['혐의 공정'])], norm(r0['결함유형(Bin)']))
    cands = by_cell.get(key, [])
    cause = norm(r0['추정 원인'])
    best, bs = None, -1.0
    for e in cands:
        if cause and (cause == e['name'] or cause in e['name']):
            best, bs = e, 1.0
            break
        t, tt = tokens(cause), tokens(e['name'])
        s = len(t & tt) / max(1, len(t | tt))
        if s > bs:
            best, bs = e, s
    if best is None:
        print(f"  !! 매칭 실패: {mid} {key} {cause[:40]}")
        continue
    link[mid] = best['id']

def day_of(r):
    d = r['발생일자']
    return str(d.date() if isinstance(d, datetime.datetime) else d)[:10]


# 집계 마감일 = 대장의 마지막 발생일. 상태 판정의 기준점이라 코드로 내보낸다.
CUTOFF = max(day_of(r) for r in recs)

# 원인별 로그 집계
agg = collections.defaultdict(lambda: {'ids': [], 'occurrences': 0, 'directional': None,
                                       'metrology': None, 'open': False, 'last_seen': ''})
for mid, cid in link.items():
    rs = by_id[mid]
    a = agg[cid]
    a['ids'].append(mid)
    a['occurrences'] += len(rs)
    d = clock_of(rs[0]['방향성 여부'])
    if d and not a['directional']:
        a['directional'] = d
    a['metrology'] = a['metrology'] or norm(rs[0]['계측 방법'])
    # 대장의 '상태' 열이 곧 상태다. 코드는 이 값을 그대로 쓴다.
    if any(norm(r['상태']) != '종결' for r in rs):
        a['open'] = True
    a['last_seen'] = max(a['last_seen'], max(day_of(r) for r in rs))

for e in entries:
    a = agg.get(e['id'])
    if a:
        e['logIds'] = sorted(a['ids'])
        e['occurrences'] = a['occurrences']
        e['metrology'] = a['metrology']
        e['lastSeen'] = a['last_seen']
        e['openCase'] = a['open']
        if a['directional']:
            e['directional'] = a['directional']
    else:
        e['logIds'], e['occurrences'], e['openCase'] = [], 0, False

linked = sum(1 for e in entries if e['logIds'])
print(f"로그와 연결된 원인 {linked}/{len(entries)}, 방향성 보유 {sum(1 for e in entries if e.get('directional'))}건")

# ─────────────────────────────────────────────── TS 출력
def ts(v, indent=0):
    pad = '  ' * indent
    if v is None:
        return 'undefined'
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, str):
        return "'" + v.replace('\\', '\\\\').replace("'", "\\'") + "'"
    if isinstance(v, list):
        if not v:
            return '[]'
        inner = ',\n'.join(pad + '  ' + ts(x, indent + 1) for x in v)
        return '[\n' + inner + ',\n' + pad + ']'
    if isinstance(v, dict):
        inner = ',\n'.join(pad + '  ' + k + ': ' + ts(x, indent + 1) for k, x in v.items())
        return '{\n' + inner + ',\n' + pad + '}'
    raise TypeError(type(v))


HEAD = """/* 자동 생성 — 직접 수정하지 말 것.
 *
 * 생성: python scripts/gen_from_xlsx.py
 * 출처: 반도체 불량 분석 개선안.xlsx · 불량 대응 log.xlsx (둘 다 읽기 전용 원본)
 * 생성 시각: %s
 *
 * 엑셀이 소스 오브 트루스다. 내용이 잘못됐으면 이 파일이 아니라 엑셀을 고치고
 * 스크립트를 다시 돌린다.
 */
import type { CauseEntry, ExclusionNote } from './causes';

""" % datetime.datetime.now().strftime('%Y-%m-%d %H:%M')

out = HEAD
out += '/** 대장의 집계 마감일 = 마지막 발생일. 상태 판정의 기준점 (domain/caseStatus.ts) */\n'
out += "export const LOG_CUTOFF = '" + CUTOFF + "';\n\n"
out += 'export const CAUSE_MATRIX: CauseEntry[] = ' + ts(entries) + ';\n\n'
out += '/** 해당 공정이 이 결함유형을 유발하지 않는다고 본 근거 — 빈칸과 "배제됨"은 다르다 */\n'
out += 'export const EXCLUSION_NOTES: ExclusionNote[] = ' + ts(exclusions) + ';\n'

dst = os.path.join(ROOT, 'src', 'domain', 'causeMatrix.generated.ts')
with open(dst, 'w', encoding='utf-8', newline='\n') as f:
    f.write(out)
print('생성:', os.path.relpath(dst, ROOT))

# ─────────────────────────────────────────────── 이력 시드
seed = []
for mid, rs in by_id.items():
    for r in rs:
        d = r['발생일자']
        d = d.date() if isinstance(d, datetime.datetime) else d
        seed.append({
            'caseId': mid,
            'causeId': link.get(mid),
            'date': str(d)[:10],
            'process': PROC_ID[norm(r['혐의 공정'])],
            'pattern': norm(r['결함유형(Bin)']),
            'lotId': norm(r['Lot ID']),
            'affected': norm(r['Lot 당 불량 Wafer 매수']),
            'symptom': norm(r['불량 현상 (측정 결과 요약)']),
            'metrology': norm(r['계측 방법']),
            'cause': norm(r['추정 원인']),
            'resolution': norm(r['해결 조치 (즉시 대응)']),
            'improvement': norm(r['개선 조치 (재발방지)']),
            'outcome': norm(r['조치 결과']),
            'team': norm(r['담당팀']),
            'owner': norm(r['담당자']),
            'closed': norm(r['상태']) == '종결',
        })
seed.sort(key=lambda x: (x['date'], x['caseId']))

HEAD2 = HEAD.replace("import type { CauseEntry, ExclusionNote } from './causes';",
                     "import type { DefectPatternId, ProcessId } from '../domain/causes';")
out2 = HEAD2 + """export interface DefectCaseRecord {
  /** 불량 대응 Log의 관리번호 (같은 번호 = 같은 결함유형의 재발 이력) */
  caseId: string;
  /** 대응하는 CAUSE_MATRIX 항목 id */
  causeId?: string;
  date: string;
  process: ProcessId;
  pattern: DefectPatternId;
  lotId: string;
  affected: string;
  symptom: string;
  metrology: string;
  cause: string;
  resolution: string;
  improvement: string;
  outcome: string;
  team: string;
  owner: string;
  /** 대장 '상태' 열 — 종결이면 true */
  closed: boolean;
}

"""
out2 += "export const LOG_CUTOFF = '" + CUTOFF + "';\n\n"
out2 += 'export const DEFECT_CASES: DefectCaseRecord[] = ' + ts(seed) + ';\n'
dst2 = os.path.join(ROOT, 'src', 'services', 'historySeed.generated.ts')
with open(dst2, 'w', encoding='utf-8', newline='\n') as f:
    f.write(out2)
print('생성:', os.path.relpath(dst2, ROOT), f'({len(seed)}건)')
